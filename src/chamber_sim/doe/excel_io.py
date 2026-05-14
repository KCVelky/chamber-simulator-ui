from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover - user environment issue
    raise ImportError(
        "Le module DOE nécessite openpyxl. Installe-le avec : pip install openpyxl"
    ) from exc

from .schemas import OUTPUT_SHEETS, OutputSheetSchema, SheetSchema




@dataclass(frozen=True)
class ExistingRunStatus:
    """Last known execution state for one run_id in runs_results."""

    run_id: str
    row_number: int
    status: str | None
    success_flag: bool | None
    input_sheet: str | None = None
    finished_at: Any | None = None
    error_message: str | None = None


def _cell_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "oui", "y"}:
        return True
    if text in {"false", "0", "no", "non", "n"}:
        return False
    return None


@dataclass(frozen=True)
class WorkbookSheetInfo:
    name: str
    max_row: int
    max_column: int


@dataclass(frozen=True)
class DOEWorkbookInfo:
    path: Path
    sheets: list[WorkbookSheetInfo]

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.name for sheet in self.sheets]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def discover_workbook(path: str | Path) -> DOEWorkbookInfo:
    """Return sheet names and dimensions without loading formulas for editing."""

    xlsx_path = Path(path).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheets = [
            WorkbookSheetInfo(name=ws.title, max_row=ws.max_row, max_column=ws.max_column)
            for ws in wb.worksheets
        ]
    finally:
        wb.close()
    return DOEWorkbookInfo(path=xlsx_path, sheets=sheets)


def read_header(path: str | Path, schema: SheetSchema) -> list[str]:
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if schema.name not in wb.sheetnames:
            raise KeyError(f"Feuille absente : {schema.name}")
        ws = wb[schema.name]
        return [_normalize_header(cell.value) for cell in ws[schema.header_row]]
    finally:
        wb.close()


def read_sheet_records(path: str | Path, schema: SheetSchema) -> list[dict[str, Any]]:
    """Read a DOE sheet into a list of dictionaries.

    Blank columns and blank rows are ignored. Headers are read from schema.header_row.
    """

    wb = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if schema.name not in wb.sheetnames:
            raise KeyError(f"Feuille absente : {schema.name}")
        ws = wb[schema.name]

        raw_headers = [_normalize_header(cell.value) for cell in ws[schema.header_row]]
        header_positions = [(idx, header) for idx, header in enumerate(raw_headers, start=1) if header]
        headers = [header for _, header in header_positions]

        records: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=schema.header_row + 1, max_row=ws.max_row):
            record: dict[str, Any] = {}
            has_data = False
            for col_idx, header in header_positions:
                value = _normalize_cell(row[col_idx - 1].value)
                record[header] = value
                if value is not None:
                    has_data = True
            if has_data:
                records.append(record)

        # Keep only columns declared by the sheet or found in the workbook.
        # This avoids polluting records with trailing empty Excel columns.
        return [{header: rec.get(header) for header in headers} for rec in records]
    finally:
        wb.close()


def copy_template_to_results(template_path: str | Path, results_path: str | Path, overwrite: bool = False) -> Path:
    src = Path(template_path).expanduser().resolve()
    dst = Path(results_path).expanduser().resolve()
    if not src.exists():
        raise FileNotFoundError(f"Template DOE introuvable : {src}")
    if dst.exists() and not overwrite:
        return dst
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst


def _apply_output_sheet_style(ws: Worksheet, n_headers: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, n_headers + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(cell.value or "")) + 2, 12), 28)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_headers)}1"


def ensure_output_sheets(results_path: str | Path, schemas: dict[str, OutputSheetSchema] | None = None) -> None:
    """Create the output sheets if absent and preserve existing data if present."""

    schemas = schemas or OUTPUT_SHEETS
    path = Path(results_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Fichier résultats introuvable : {path}")

    wb = load_workbook(path)
    try:
        for sheet_schema in schemas.values():
            if sheet_schema.name in wb.sheetnames:
                ws = wb[sheet_schema.name]
                if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
                    ws.append(list(sheet_schema.headers))
                    _apply_output_sheet_style(ws, len(sheet_schema.headers))
                continue

            ws = wb.create_sheet(sheet_schema.name)
            ws.append(list(sheet_schema.headers))
            _apply_output_sheet_style(ws, len(sheet_schema.headers))
        wb.save(path)
    finally:
        wb.close()


def prepare_results_workbook(
    template_path: str | Path,
    results_path: str | Path | None = None,
    overwrite: bool = False,
) -> Path:
    """Copy the DOE template and ensure standard output sheets exist."""

    template = Path(template_path).expanduser().resolve()
    if results_path is None:
        results_path = template.with_name(f"{template.stem}_results{template.suffix}")
    results = copy_template_to_results(template, results_path, overwrite=overwrite)
    ensure_output_sheets(results)
    return results


def get_existing_run_ids(results_path: str | Path, sheet_name: str = "runs_results") -> set[str]:
    """Return completed run IDs already present in a results sheet.

    This will be used in later steps for resume/skip logic.
    """

    path = Path(results_path).expanduser().resolve()
    if not path.exists():
        return set()

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return set()
        ws = wb[sheet_name]
        headers = [_normalize_header(cell.value) for cell in ws[1]]
        try:
            run_id_idx = headers.index("run_id") + 1
        except ValueError:
            return set()

        ids: set[str] = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            value = _normalize_cell(row[run_id_idx - 1].value)
            if value is not None:
                ids.add(str(value))
        return ids
    finally:
        wb.close()



def get_run_statuses(results_path: str | Path, sheet_name: str = "runs_results") -> dict[str, ExistingRunStatus]:
    """Return the last known status for every run_id in the results workbook.

    The last row wins. This is useful if the user manually edited or appended rows.
    Resume logic should skip only rows whose last status has ``success_flag=True``.
    """

    path = Path(results_path).expanduser().resolve()
    if not path.exists():
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        headers = [_normalize_header(cell.value) for cell in ws[1]]
        index = {header: idx for idx, header in enumerate(headers) if header}
        if "run_id" not in index:
            return {}

        out: dict[str, ExistingRunStatus] = {}
        for row_number, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            raw_run_id = _normalize_cell(row[index["run_id"]].value)
            if raw_run_id is None:
                continue
            run_id = str(raw_run_id).strip()
            if not run_id:
                continue

            def get_value(name: str) -> Any:
                col = index.get(name)
                return None if col is None or col >= len(row) else _normalize_cell(row[col].value)

            out[run_id] = ExistingRunStatus(
                run_id=run_id,
                row_number=row_number,
                status=str(get_value("status")) if get_value("status") is not None else None,
                success_flag=_cell_bool(get_value("success_flag")),
                input_sheet=str(get_value("input_sheet")) if get_value("input_sheet") is not None else None,
                finished_at=get_value("finished_at"),
                error_message=str(get_value("error_message")) if get_value("error_message") is not None else None,
            )
        return out
    finally:
        wb.close()

def count_records(records: Iterable[dict[str, Any]]) -> int:
    return sum(1 for _ in records)


def _header_index_map(ws: Worksheet) -> dict[str, int]:
    return {_normalize_header(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if _normalize_header(cell.value)}


def _delete_rows_for_run_id(wb, run_id: str, sheet_names: Iterable[str]) -> None:
    if not run_id:
        return
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = _header_index_map(ws)
        run_col = headers.get("run_id")
        if run_col is None:
            continue
        for row_idx in range(ws.max_row, 1, -1):
            value = _normalize_cell(ws.cell(row=row_idx, column=run_col).value)
            if value is not None and str(value).strip() == run_id:
                ws.delete_rows(row_idx, 1)


def _append_dict_rows(ws: Worksheet, rows: Iterable[dict[str, Any]], headers: list[str]) -> None:
    for row in rows:
        ws.append([row.get(header) for header in headers])


def append_output_rows(
    results_path: str | Path,
    rows_by_sheet: dict[str, Iterable[Mapping[str, Any]]],
    *,
    replace_run_id: str | None = None,
) -> None:
    """Append dictionaries to DOE output sheets using their declared headers.

    Unknown keys are ignored, missing keys are written as blank cells. If replace_run_id
    is provided, previous rows with the same run_id are removed from every target sheet
    before appending. This keeps manual re-runs clean during development.
    """

    path = Path(results_path).expanduser().resolve()
    ensure_output_sheets(path)

    wb = load_workbook(path)
    try:
        if replace_run_id:
            _delete_rows_for_run_id(wb, replace_run_id, rows_by_sheet.keys())

        for sheet_name, rows_iter in rows_by_sheet.items():
            rows = [dict(row) for row in rows_iter]
            if not rows:
                continue
            if sheet_name not in OUTPUT_SHEETS:
                raise KeyError(f"Feuille de sortie inconnue : {sheet_name}")
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(list(OUTPUT_SHEETS[sheet_name].headers))
                _apply_output_sheet_style(ws, len(OUTPUT_SHEETS[sheet_name].headers))
            else:
                ws = wb[sheet_name]

            headers = list(OUTPUT_SHEETS[sheet_name].headers)
            existing_headers = [_normalize_header(cell.value) for cell in ws[1]][: len(headers)]
            if existing_headers != headers:
                # Keep the sheet usable even if a user edited the first row.
                ws.delete_rows(1, 1)
                ws.insert_rows(1, 1)
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx).value = header
                _apply_output_sheet_style(ws, len(headers))

            _append_dict_rows(ws, rows, headers)

        wb.save(path)
    finally:
        wb.close()
