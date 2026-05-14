from __future__ import annotations

import math
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_io import ensure_output_sheets, read_sheet_records
from .schemas import DOE_RUN_SHEETS, OUTPUT_SHEETS


@dataclass(frozen=True)
class ScoringPolicy:
    """Robust score used to rank DOE candidates.

    Smaller is better. The score combines precision, robustness, reliability and
    compute time. The defaults intentionally penalize failures strongly so a
    method that is sometimes excellent but often unstable is not selected.
    """

    p90_weight: float = 0.25
    failure_penalty: float = 10.0
    time_weight: float = 0.001
    min_success_rate: float = 0.75

    def score(self, rmse_3d: float | None, p90_err_3d: float | None, success_rate: float | None, mean_time_s: float | None) -> float:
        if rmse_3d is None or p90_err_3d is None or success_rate is None:
            return 1e9
        penalty = self.failure_penalty * max(0.0, 1.0 - float(success_rate))
        if float(success_rate) < self.min_success_rate:
            penalty += self.failure_penalty
        return float(rmse_3d) + self.p90_weight * float(p90_err_3d) + penalty + self.time_weight * float(mean_time_s or 0.0)


@dataclass(frozen=True)
class CandidateScore:
    candidate_type: str
    candidate_id: str
    rank: int
    n_runs: int
    n_success: int
    success_rate: float
    rmse_3d: float | None
    median_err_3d: float | None
    p90_err_3d: float | None
    mean_time_s: float | None
    score: float
    metadata: dict[str, Any] = field(default_factory=dict)

    def as_row(self, stage: str, input_sheet: str, notes: str = "") -> dict[str, Any]:
        return {
            "stage": stage,
            "input_sheet": input_sheet,
            "candidate_type": self.candidate_type,
            "candidate_id": self.candidate_id,
            "rank": self.rank,
            "n_runs": self.n_runs,
            "n_success": self.n_success,
            "success_rate": self.success_rate,
            "rmse_3d": self.rmse_3d,
            "median_err_3d": self.median_err_3d,
            "p90_err_3d": self.p90_err_3d,
            "mean_time_s": self.mean_time_s,
            "score": self.score,
            "notes": notes,
        }


@dataclass(frozen=True)
class AutomationSelections:
    geometry_slots: dict[str, str] = field(default_factory=dict)
    algorithm_slots: dict[str, dict[str, Any]] = field(default_factory=dict)
    decisions: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class DOEAnalysisReport:
    selections: AutomationSelections
    stage_summary_rows: list[dict[str, Any]] = field(default_factory=list)
    candidate_rows: list[dict[str, Any]] = field(default_factory=list)
    decision_rows: list[dict[str, Any]] = field(default_factory=list)
    factor_effect_rows: list[dict[str, Any]] = field(default_factory=list)


_NUMERIC_EMPTY = {"", "none", "null", "nan", "na", "n/a", "-"}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and value.strip().lower() in _NUMERIC_EMPTY:
        return None
    try:
        out = float(value)
    except Exception:
        return None
    if math.isnan(out) or math.isinf(out):
        return None
    return out


def _to_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "oui", "y"}:
        return True
    if text in {"false", "0", "no", "non", "n"}:
        return False
    return None


def _mean(values: Iterable[float | None]) -> float | None:
    vals = [float(v) for v in values if v is not None and not math.isnan(float(v))]
    if not vals:
        return None
    return float(sum(vals) / len(vals))


def _rmse(values: Iterable[float | None]) -> float | None:
    vals = [float(v) for v in values if v is not None and not math.isnan(float(v))]
    if not vals:
        return None
    return float(math.sqrt(sum(v * v for v in vals) / len(vals)))


def _percentile(values: Iterable[float | None], q: float) -> float | None:
    vals = sorted(float(v) for v in values if v is not None and not math.isnan(float(v)))
    if not vals:
        return None
    if len(vals) == 1:
        return float(vals[0])
    q = min(max(float(q), 0.0), 1.0)
    pos = q * (len(vals) - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return float(vals[lo])
    frac = pos - lo
    return float(vals[lo] * (1.0 - frac) + vals[hi] * frac)


def _timestamp() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def read_output_rows(results_path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    path = Path(results_path).expanduser().resolve()
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [_normalize_header(cell.value) for cell in ws[1]]
        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            rec = {h: _normalize_cell(row[i].value) for i, h in enumerate(headers) if h}
            if any(value is not None for value in rec.values()):
                out.append(rec)
        return out
    finally:
        wb.close()


def _apply_table_style(ws, n_headers: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for idx in range(1, n_headers + 1):
        c = ws.cell(row=1, column=idx)
        c.fill = fill
        c.font = font
        ws.column_dimensions[get_column_letter(idx)].width = min(max(len(str(c.value or "")) + 2, 12), 34)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_headers)}1"


def replace_output_sheet_rows(results_path: str | Path, sheet_name: str, rows: Sequence[Mapping[str, Any]]) -> None:
    """Replace the data section of one standard output sheet.

    This is used for automated summaries. Run-level results are still appended by
    the existing result writer after each simulation.
    """

    if sheet_name not in OUTPUT_SHEETS:
        raise KeyError(f"Feuille de sortie inconnue : {sheet_name}")
    path = Path(results_path).expanduser().resolve()
    ensure_output_sheets(path)
    wb = load_workbook(path)
    try:
        schema = OUTPUT_SHEETS[sheet_name]
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)
        headers = list(schema.headers)
        ws.append(headers)
        _apply_table_style(ws, len(headers))
        for row in rows:
            ws.append([row.get(h) for h in headers])
        wb.save(path)
    finally:
        wb.close()


def _filter_latest_rows_by_run(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for row in rows:
        run_id = str(row.get("run_id") or "").strip()
        if not run_id:
            continue
        latest[run_id] = dict(row)
    return list(latest.values())


def _rows_for_input_sheet(results_path: str | Path, input_sheet: str) -> list[dict[str, Any]]:
    rows = _filter_latest_rows_by_run(read_output_rows(results_path, "runs_results"))
    return [row for row in rows if str(row.get("input_sheet") or "").strip() == input_sheet]


def _aggregate_candidate(candidate_type: str, candidate_id: str, rows: Sequence[Mapping[str, Any]], policy: ScoringPolicy, rank: int = 0) -> CandidateScore:
    n_runs = len(rows)
    successes = [row for row in rows if _to_bool(row.get("success_flag")) is True]
    n_success = len(successes)
    success_rate = 0.0 if n_runs == 0 else n_success / n_runs
    success_errors = [_to_float(row.get("err_3d")) for row in successes]
    rmse = _rmse(success_errors)
    med = _percentile(success_errors, 0.50)
    p90 = _percentile(success_errors, 0.90)
    tmean = _mean(_to_float(row.get("time_s")) for row in rows)
    score = policy.score(rmse, p90, success_rate, tmean)
    return CandidateScore(
        candidate_type=candidate_type,
        candidate_id=str(candidate_id),
        rank=rank,
        n_runs=n_runs,
        n_success=n_success,
        success_rate=success_rate,
        rmse_3d=rmse,
        median_err_3d=med,
        p90_err_3d=p90,
        mean_time_s=tmean,
        score=score,
    )


def rank_candidates(
    rows: Sequence[Mapping[str, Any]],
    *,
    group_key: str,
    candidate_type: str,
    policy: ScoringPolicy,
) -> list[CandidateScore]:
    groups: dict[str, list[Mapping[str, Any]]] = {}
    for row in rows:
        key = str(row.get(group_key) or "").strip()
        if not key:
            continue
        groups.setdefault(key, []).append(row)

    candidates = [_aggregate_candidate(candidate_type, key, group_rows, policy) for key, group_rows in groups.items()]
    candidates.sort(key=lambda item: (item.score, -item.success_rate, item.rmse_3d if item.rmse_3d is not None else 1e9, item.candidate_id))
    return [CandidateScore(**{**candidate.__dict__, "rank": idx}) for idx, candidate in enumerate(candidates, start=1)]


def summarize_sheet(results_path: str | Path, input_sheet: str, policy: ScoringPolicy) -> dict[str, Any] | None:
    rows = _rows_for_input_sheet(results_path, input_sheet)
    if not rows:
        return None
    summary = _aggregate_candidate("sheet", input_sheet, rows, policy, rank=1)
    best_rows = sorted(rows, key=lambda row: _to_float(row.get("score")) if _to_float(row.get("score")) is not None else 1e9)
    best_row = best_rows[0] if best_rows else {}
    return {
        "stage": DOE_RUN_SHEETS[input_sheet].stage if input_sheet in DOE_RUN_SHEETS else None,
        "input_sheet": input_sheet,
        "n_runs": summary.n_runs,
        "n_success": summary.n_success,
        "success_rate": summary.success_rate,
        "rmse_3d": summary.rmse_3d,
        "median_err_3d": summary.median_err_3d,
        "p90_err_3d": summary.p90_err_3d,
        "mean_time_s": summary.mean_time_s,
        "best_run_id": best_row.get("run_id"),
        "best_score": best_row.get("score"),
        "selected_geometry_or_algo": None,
        "notes": None,
    }


def _metadata_by_run_id(workbook_path: str | Path, sheet_names: Iterable[str]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for sheet_name in sheet_names:
        if sheet_name not in DOE_RUN_SHEETS:
            continue
        for rec in read_sheet_records(workbook_path, DOE_RUN_SHEETS[sheet_name]):
            run_id = str(rec.get("run_id") or rec.get("pb_run") or rec.get("alpha_case") or "").strip()
            if run_id:
                out[run_id] = dict(rec)
    return out


def _algo_signature(meta: Mapping[str, Any]) -> str:
    keys = [
        "model_type",
        "init_method",
        "estimate_alpha",
        "alpha_init",
        "alpha_grid_size",
        "lam",
        "max_iter",
        "fd_eps",
    ]
    parts = []
    for key in keys:
        value = meta.get(key)
        if value is not None:
            parts.append(f"{key}={value}")
    return "; ".join(parts) if parts else "baseline"


def _algorithm_settings_from_meta(meta: Mapping[str, Any]) -> dict[str, Any]:
    mapping = {
        "model_type": "mleem_model_type",
        "init_method": "mleem_init_method",
        "estimate_alpha": "mleem_estimate_alpha",
        "lam": "mleem_lam",
        "max_iter": "mleem_max_iter",
        "fd_eps": "mleem_fd_eps",
        "alpha_init": "mleem_alpha_init",
        "alpha_grid_size": "mleem_alpha_grid_size",
    }
    out: dict[str, Any] = {}
    for source_key, target_key in mapping.items():
        if meta.get(source_key) is not None:
            out[target_key] = meta.get(source_key)
    return out


def rank_algorithm_candidates(workbook_path: str | Path, results_path: str | Path, policy: ScoringPolicy) -> tuple[list[CandidateScore], dict[str, dict[str, Any]]]:
    rows = []
    for sheet_name in ("C2_alpha_refine", "C1_PB12_common"):
        rows.extend(_rows_for_input_sheet(results_path, sheet_name))
    meta_by_run = _metadata_by_run_id(workbook_path, ("C1_PB12_common", "C2_alpha_refine"))

    grouped: dict[str, list[Mapping[str, Any]]] = {}
    settings_by_signature: dict[str, dict[str, Any]] = {}
    for row in rows:
        run_id = str(row.get("run_id") or "").strip()
        meta = meta_by_run.get(run_id, {})
        signature = _algo_signature(meta)
        grouped.setdefault(signature, []).append(row)
        if signature not in settings_by_signature:
            settings_by_signature[signature] = _algorithm_settings_from_meta(meta)

    candidates = [_aggregate_candidate("algorithm", signature, group_rows, policy) for signature, group_rows in grouped.items()]
    candidates.sort(key=lambda item: (item.score, -item.success_rate, item.rmse_3d if item.rmse_3d is not None else 1e9, item.candidate_id))
    ranked = [CandidateScore(**{**candidate.__dict__, "rank": idx}) for idx, candidate in enumerate(candidates, start=1)]
    return ranked, settings_by_signature


def compute_factor_effects(results_path: str | Path, workbook_path: str | Path, policy: ScoringPolicy) -> list[dict[str, Any]]:
    rows = _rows_for_input_sheet(results_path, "B_PB12")
    if not rows:
        return []
    meta_by_run = _metadata_by_run_id(workbook_path, ("B_PB12",))
    factors = ["B1_snr", "B2_duration_s", "B3_window_s", "B4_hop_ratio", "B5_trim_frac", "B6_gain_err_db", "B7_mic_pos_err_mm", "B8_c_err_pct"]
    out: list[dict[str, Any]] = []
    for factor in factors:
        groups: dict[str, list[Mapping[str, Any]]] = {}
        for row in rows:
            meta = meta_by_run.get(str(row.get("run_id") or ""), {})
            value = meta.get(factor)
            if value is None:
                continue
            groups.setdefault(str(value), []).append(row)
        if len(groups) < 2:
            continue
        scored = []
        for level, group_rows in groups.items():
            candidate = _aggregate_candidate("factor_level", level, group_rows, policy)
            scored.append((level, candidate))
        scored.sort(key=lambda item: item[1].score)
        best_level, best = scored[0]
        worst_level, worst = scored[-1]
        best_err = best.rmse_3d if best.rmse_3d is not None else 1e9
        worst_err = worst.rmse_3d if worst.rmse_3d is not None else 1e9
        out.append(
            {
                "stage": "B",
                "factor": factor,
                "best_level": best_level,
                "worst_level": worst_level,
                "effect_on_score": float(worst.score - best.score),
                "effect_on_rmse_3d": None if best_err >= 1e9 or worst_err >= 1e9 else float(worst_err - best_err),
                "best_score": best.score,
                "worst_score": worst.score,
                "notes": "Plus l'effet est élevé, plus le facteur semble influent dans les runs disponibles.",
            }
        )
    out.sort(key=lambda row: abs(float(row.get("effect_on_score") or 0.0)), reverse=True)
    return out


class DOEAnalyzer:
    """Read DOE result sheets, rank candidates and persist decisions."""

    def __init__(self, workbook_path: str | Path, results_path: str | Path, policy: ScoringPolicy | None = None) -> None:
        self.workbook_path = Path(workbook_path).expanduser().resolve()
        self.results_path = Path(results_path).expanduser().resolve()
        self.policy = policy or ScoringPolicy()

    def rows_for_sheet(self, input_sheet: str) -> list[dict[str, Any]]:
        return _rows_for_input_sheet(self.results_path, input_sheet)

    def stage_summary_rows(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for sheet in DOE_RUN_SHEETS:
            summary = summarize_sheet(self.results_path, sheet, self.policy)
            if summary is not None:
                rows.append(summary)
        return rows

    def select_geometries_after_screen(self, top_n: int = 2) -> list[CandidateScore]:
        rows = self.rows_for_sheet("A_runs_screen")
        return rank_candidates(rows, group_key="geom_run_id", candidate_type="geometry", policy=self.policy)[:top_n]

    def select_geometries_after_map(self, top_n: int = 2) -> list[CandidateScore]:
        rows = self.rows_for_sheet("A_runs_map")
        candidates = rank_candidates(rows, group_key="geom_run_id", candidate_type="geometry", policy=self.policy)
        if not candidates:
            candidates = self.select_geometries_after_screen(top_n=top_n)
        return candidates[:top_n]

    def select_algorithm(self) -> tuple[CandidateScore | None, dict[str, Any]]:
        ranked, settings = rank_algorithm_candidates(self.workbook_path, self.results_path, self.policy)
        if not ranked:
            return None, {}
        best = ranked[0]
        return best, settings.get(best.candidate_id, {})

    def build_report(self) -> DOEAnalysisReport:
        geometry_slots: dict[str, str] = {}
        algorithm_slots: dict[str, dict[str, Any]] = {}
        decisions: dict[str, Any] = {}
        candidate_rows: list[dict[str, Any]] = []
        decision_rows: list[dict[str, Any]] = []
        timestamp = _timestamp()

        screen_best = self.select_geometries_after_screen(top_n=2)
        for idx, candidate in enumerate(screen_best, start=1):
            slot = f"BEST_A_{idx}"
            geometry_slots[slot] = candidate.candidate_id
            decisions[slot] = candidate.candidate_id
            candidate_rows.append(candidate.as_row("A", "A_runs_screen", notes=f"Sélection automatique {slot}"))
            decision_rows.append(
                {
                    "decision_key": slot,
                    "selected_value": candidate.candidate_id,
                    "source_stage": "A_screen",
                    "score": candidate.score,
                    "n_runs": candidate.n_runs,
                    "timestamp": timestamp,
                    "notes": "Meilleure géométrie issue du screening A_runs_screen.",
                }
            )
        if len(screen_best) == 1:
            # In test/limited mode only one geometry may be available. Reuse it so
            # downstream sheets using BEST_A_2 can still be dry-run or smoke-tested.
            geometry_slots["BEST_A_2"] = screen_best[0].candidate_id
            decisions["BEST_A_2"] = screen_best[0].candidate_id
            decision_rows.append(
                {
                    "decision_key": "BEST_A_2",
                    "selected_value": screen_best[0].candidate_id,
                    "source_stage": "A_screen",
                    "score": screen_best[0].score,
                    "n_runs": screen_best[0].n_runs,
                    "timestamp": timestamp,
                    "notes": "Fallback automatique car une seule géométrie est disponible dans les résultats actuels.",
                }
            )

        map_best = self.select_geometries_after_map(top_n=2)
        for candidate in map_best:
            candidate_rows.append(candidate.as_row("A", "A_runs_map", notes="Classement après cartographie complète"))
        if map_best:
            geometry_slots["BEST_GLOBAL"] = map_best[0].candidate_id
            geometry_slots["BEST_1"] = map_best[0].candidate_id
            decisions["BEST_GLOBAL"] = map_best[0].candidate_id
            decisions["BEST_1"] = map_best[0].candidate_id
            decision_rows.append(
                {
                    "decision_key": "BEST_GLOBAL",
                    "selected_value": map_best[0].candidate_id,
                    "source_stage": "A_map" if self.rows_for_sheet("A_runs_map") else "A_screen",
                    "score": map_best[0].score,
                    "n_runs": map_best[0].n_runs,
                    "timestamp": timestamp,
                    "notes": "Géométrie globale retenue pour les stages B/C/D.",
                }
            )
            decision_rows.append(
                {
                    "decision_key": "BEST_1",
                    "selected_value": map_best[0].candidate_id,
                    "source_stage": "A_map" if self.rows_for_sheet("A_runs_map") else "A_screen",
                    "score": map_best[0].score,
                    "n_runs": map_best[0].n_runs,
                    "timestamp": timestamp,
                    "notes": "Géométrie 1 pour validation finale D.",
                }
            )
        if len(map_best) >= 2:
            geometry_slots["BEST_2"] = map_best[1].candidate_id
            decisions["BEST_2"] = map_best[1].candidate_id
            decision_rows.append(
                {
                    "decision_key": "BEST_2",
                    "selected_value": map_best[1].candidate_id,
                    "source_stage": "A_map" if self.rows_for_sheet("A_runs_map") else "A_screen",
                    "score": map_best[1].score,
                    "n_runs": map_best[1].n_runs,
                    "timestamp": timestamp,
                    "notes": "Géométrie 2 pour validation finale D.",
                }
            )
        elif map_best:
            geometry_slots["BEST_2"] = map_best[0].candidate_id
            decisions["BEST_2"] = map_best[0].candidate_id

        # Add full algorithm ranking rows.
        alg_ranked, settings_by_signature = rank_algorithm_candidates(self.workbook_path, self.results_path, self.policy)
        for candidate in alg_ranked:
            candidate_rows.append(candidate.as_row("C", "C1_C2", notes="Classement des réglages algorithmiques"))
        best_algo, best_settings = self.select_algorithm()
        if best_algo is not None:
            algorithm_slots["BEST_GLOBAL"] = best_settings
            decisions["BEST_ALGO"] = best_algo.candidate_id
            decision_rows.append(
                {
                    "decision_key": "BEST_ALGO",
                    "selected_value": best_algo.candidate_id,
                    "source_stage": "C2/C1",
                    "score": best_algo.score,
                    "n_runs": best_algo.n_runs,
                    "timestamp": timestamp,
                    "notes": f"Réglages appliqués à D via algo_setting=BEST_GLOBAL : {best_settings}",
                }
            )

        stage_summary = self.stage_summary_rows()
        # Fill selected value in stage_summary when possible.
        for row in stage_summary:
            if row.get("input_sheet") in {"A_runs_screen", "A_runs_map"}:
                row["selected_geometry_or_algo"] = decisions.get("BEST_GLOBAL") or decisions.get("BEST_A_1")
            elif row.get("input_sheet") in {"C1_PB12_common", "C2_alpha_refine"}:
                row["selected_geometry_or_algo"] = decisions.get("BEST_ALGO")

        factor_rows = compute_factor_effects(self.results_path, self.workbook_path, self.policy)

        return DOEAnalysisReport(
            selections=AutomationSelections(geometry_slots=geometry_slots, algorithm_slots=algorithm_slots, decisions=decisions),
            stage_summary_rows=stage_summary,
            candidate_rows=candidate_rows,
            decision_rows=decision_rows,
            factor_effect_rows=factor_rows,
        )

    def write_report(self, report: DOEAnalysisReport | None = None) -> DOEAnalysisReport:
        report = report or self.build_report()
        replace_output_sheet_rows(self.results_path, "stage_summary", report.stage_summary_rows)
        replace_output_sheet_rows(self.results_path, "best_candidates", report.candidate_rows)
        replace_output_sheet_rows(self.results_path, "automation_decisions", report.decision_rows)
        replace_output_sheet_rows(self.results_path, "factor_effects", report.factor_effect_rows)
        return report
